from openpyxl import load_workbook

try:
    wb = load_workbook('fiches_stages.xlsx')
    ws = wb.active
    print("Noms dans le template:")
    for i in range(9, 37):
        val = ws[f'B{i}'].value
        print(f'Row {i}: {val}')
except FileNotFoundError:
    print("Fichier fiches_stages.xlsx n'existe pas!")
except Exception as e:
    print(f"Erreur: {e}")
