from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Créer un nouveau workbook
wb = Workbook()
ws = wb.active
ws.title = "Fiches de Stage"

# Liste des 27 étudiants
etudiants = [
    "AMANI KOUASSI STEVEN",
    "ASSALE N'DAH JEAN",
    "BLE GAYE MARC DAVID",
    "COULIBALY NAHOUO ALBERT",
    "DEMBELE MADOUSSOU EUNICE",
    "EZIN ASSOUHAN DERIC STEPHANE",
    "FOFANA ISSOUF",
    "GOLITI BI MARC ANI KEVIN",
    "KAMATE FLAGNAN",
    "KANI EHOUMAN MATTHIAS",
    "KANTE YOUSSOUF AZIZ",
    "KOFFI ASSABA LINDA SEPHORA",
    "KONAN KONAN N'DRI ROMUALD",
    "KONE GNOUWETCHA ANGE",
    "KOUADIO AKOUA MATHANIA",
    "KOUAKOUSSUI YANN EZECHIEL AYMARD",
    "KOUAME SOURALEH JATHE",
    "LOBA EMMANUELLA VALENCIA",
    "M'BO MELVIN MARC EMMANUEL",
    "OFFO ANGE EMMANUEL",
    "OULE DESIRE-MARIE",
    "SORO FANNY AICHA",
    "SORO FOUNGNIGUE IVAN",
    "TOKAPIEU OUANGUI EZECHIEL",
    "YEBOUE KOUASSI EMMANUEL",
    "YEO YOMOYAHA",
    "ZIAO KOLO"
]

# Définir les styles
header_fill = PatternFill(start_color="0D1B3E", end_color="0D1B3E", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Ligne 9: En-têtes
headers = ["ORDRE", "NOM & PRENOMS", "CONTACT ETUDIANT", "NIVEAU", "NOM ENTREPRISE", "OUI", "NON", "SITUATION ENTREPRISE", "CONTACT ENT.", "DATE DEBUT", "DATE FIN"]
header_cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]

for col, header in zip(header_cols, headers):
    cell = ws[f'{col}9']
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = border

# Remplir les noms dans les lignes 10-36
for idx, etudiant in enumerate(etudiants, start=10):
    # Colonne A: ORDRE
    ws[f'A{idx}'].value = idx - 9
    # Colonne B: NOM
    ws[f'B{idx}'].value = etudiant
    
    # Appliquer les bordures à toutes les colonnes
    for col in header_cols:
        ws[f'{col}{idx}'].border = border
        ws[f'{col}{idx}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Ajuster les largeurs des colonnes
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 8
ws.column_dimensions['H'].width = 20
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 12
ws.column_dimensions['K'].width = 12

# Geler la ligne d'en-têtes et la colonne des noms
ws.freeze_panes = "C10"

# Sauvegarder le fichier
wb.save('fiches_stages.xlsx')
print("✅ Fichier Excel template créé avec succès!")
print(f"📊 {len(etudiants)} étudiants pré-remplis dans les lignes 10-36")
