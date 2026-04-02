from openpyxl import load_workbook
from datetime import datetime
import os
from pathlib import Path
from config import EXCEL_TEMPLATE_PATH

def format_date(date_obj):
    """Formate une date pour l'affichage"""
    if date_obj is None:
        return ""
    return date_obj.strftime("%d/%m/%Y")

def generate_excel(fiches_stages, filepath):
    """
    Remplissage du fichier Excel template avec les fiches de stage
    
    Structure du template:
    - Ligne 9: En-têtes
    - Lignes 10-36: Noms d'étudiants (27 noms fixes)
    - À chaque nom correspond une ligne où on remplit les données
    
    Colonnes:
    - A: ORDRE
    - B: NOM & PRENOMS (déjà rempli)
    - C: CONTACT ETUDIANT
    - E: NOM ENTREPRISE
    - F: OUI (stage?)
    - G: NON (pas de stage)
    - H: SITUATION ENTREPRISE (Commune)
    - I: CONTACT ENTREPRISE
    - J: DATE DEBUT
    - K: DATE FIN
    """
    
    # Chemin du template depuis config
    template_path = EXCEL_TEMPLATE_PATH
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template Excel non trouvé: {template_path}")
    
    # Charger le workbook template
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Les noms sont aux lignes 10 à 36 (27 noms)
    # On crée un dictionnaire: nom -> numéro de ligne
    noms_lignes = {}
    for row in range(10, 37):  # Lignes 10 à 36
        nom_cell = ws[f'B{row}'].value
        if nom_cell:
            # Normaliser le nom (majuscules, sans espaces superflus)
            noms_lignes[str(nom_cell).strip().upper()] = row
    
    # Remplir les données pour chaque fiche
    for fiche in fiches_stages:
        # Récupérer le nom de l'étudiant depuis la relation
        nom_etudiant = fiche.etudiant.nom if fiche.etudiant else None
        if not nom_etudiant:
            print(f"⚠️ Pas d'étudiant associé à la fiche {fiche.id}")
            continue
        
        # Chercher la ligne correspondant au nom
        nom_search = nom_etudiant.strip().upper()
        
        if nom_search not in noms_lignes:
            print(f"⚠️ Nom non trouvé dans le template: {nom_etudiant}")
            continue
        
        row = noms_lignes[nom_search]
        
        # Remplir les colonnes pour cette ligne
        
        # Colonne C: CONTACT ETUDIANT
        ws[f'C{row}'] = fiche.contact_etudiant or ""
        
        # Colonne E: NOM ENTREPRISE
        ws[f'E{row}'] = fiche.nom_entreprise or ""
        
        # Colonne F: OUI (si en stage)
        ws[f'F{row}'] = "OUI" if fiche.en_stage == "OUI" else ""
        
        # Colonne G: NON (si pas en stage)
        ws[f'G{row}'] = "NON" if fiche.en_stage == "NON" else ""
        
        # Colonne H: SITUATION ENTREPRISE (Commune)
        ws[f'H{row}'] = fiche.situation_entreprise or ""
        
        # Colonne I: CONTACT ENTREPRISE
        ws[f'I{row}'] = fiche.contact_entreprise or ""
        
        # Colonne J: DATE DEBUT
        if fiche.date_debut_stage:
            ws[f'J{row}'] = fiche.date_debut_stage.strftime("%d/%m/%Y")
        
        # Colonne K: DATE FIN (Facultative)
        if fiche.date_fin_stage:
            ws[f'K{row}'] = fiche.date_fin_stage.strftime("%d/%m/%Y")
    
    # Sauvegarder le fichier
    os.makedirs(os.path.dirname(filepath) if os.path.dirname(filepath) else ".", exist_ok=True)
    wb.save(filepath)
    
    return filepath

def update_excel(fiches_stages, filepath):
    """Met à jour le fichier Excel (alias pour plus de clarté)"""
    return generate_excel(fiches_stages, filepath)
